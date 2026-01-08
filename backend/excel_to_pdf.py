# """
# excel_to_rag.py
# ----------------

# Classifies Excel sheets and exports them optimally for RAG ingestion:

# - Detects "forms" → form-like PDF
# - Detects "structured tables" → OCR-friendly table PDF
# - Detects "multi-table / irregular blocks" → multiple PDFs
# - Detects "schema / sparse / meta" → CSV
# - Detects "large data" → CSV

# Supports both .xls and .xlsx inputs
# """

# from pathlib import Path
# import pandas as pd
# from reportlab.lib.pagesizes import A4, landscape, portrait
# from reportlab.lib import colors
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from reportlab.lib.units import mm
# import numpy as np

# # ---------------------------------------------------------------------
# # 🧩 Utility functions
# # ---------------------------------------------------------------------

# def safe_str(x):
#     return "" if pd.isna(x) else str(x).strip()

# def mm_to_pt(mm_val):
#     return mm_val * 2.83465

# # ---------------------------------------------------------------------
# # 📏 PDF Helpers
# # ---------------------------------------------------------------------

# def compute_column_widths_auto(df: pd.DataFrame, page_width_pt: float, margin_pt: float):
#     """Compute relative column widths that fit page width."""
#     usable_width = page_width_pt - 2 * margin_pt
#     ncols = len(df.columns)
#     if ncols == 0:
#         return []
#     weights = []
#     for col in df.columns:
#         col_len = len(str(col))
#         avg_len = min(df[col].astype(str).map(len).mean(), 80)
#         weights.append(max(col_len * 1.5, avg_len))
#     total_weight = sum(weights)
#     widths = [usable_width * (w / total_weight) for w in weights]
#     min_width_pt, max_width_pt = 20 * mm, 90 * mm
#     widths = [min(max(w, min_width_pt), max_width_pt) for w in widths]
#     total = sum(widths)
#     if total > usable_width:
#         scale = usable_width / total
#         widths = [w * scale for w in widths]
#     return widths

# # ---------------------------------------------------------------------
# # 🧾 PDF Table Renderer
# # ---------------------------------------------------------------------

# def df_to_pdf(df: pd.DataFrame, pdf_path: Path, title=None, orientation="auto"):
#     pdf_path.parent.mkdir(parents=True, exist_ok=True)
#     df = df.fillna("")
#     ncols = len(df.columns)
#     max_cell_len = df.astype(str).applymap(len).max().max()
#     if orientation == "auto":
#         orientation = "landscape" if ncols > 8 or max_cell_len > 40 else "portrait"
#     page_size = landscape(A4) if orientation == "landscape" else portrait(A4)
#     page_width, _ = page_size
#     margin_pt = mm_to_pt(15)

#     styles = getSampleStyleSheet()
#     normal_style = ParagraphStyle(
#         "wrapped",
#         parent=styles["Normal"],
#         fontName="Helvetica",
#         fontSize=8,
#         leading=10,
#         alignment=0,
#     )

#     doc = SimpleDocTemplate(
#         str(pdf_path),
#         pagesize=page_size,
#         leftMargin=margin_pt,
#         rightMargin=margin_pt,
#         topMargin=mm_to_pt(15),
#         bottomMargin=mm_to_pt(15),
#     )

#     data = [[Paragraph(f"<b>{safe_str(c)}</b>", normal_style) for c in df.columns]]
#     for _, row in df.iterrows():
#         data.append([Paragraph(safe_str(v).replace("\n", "<br/>"), normal_style) for v in row])

#     col_widths = compute_column_widths_auto(df, page_width, margin_pt)
#     table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
#     table.setStyle(TableStyle([
#         ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
#         ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
#         ("ALIGN", (0, 0), (-1, -1), "LEFT"),
#         ("VALIGN", (0, 0), (-1, -1), "TOP"),
#         ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
#         ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
#         ("FONTSIZE", (0, 0), (-1, -1), 8),
#     ]))

#     elements = []
#     if title:
#         elements.append(Paragraph(f"<b>{title}</b>", styles["Heading2"]))
#         elements.append(Spacer(1, 8))
#     elements.append(table)
#     doc.build(elements)
#     print(f"✅ PDF created: {pdf_path}")

# # ---------------------------------------------------------------------
# # 📊 Classification Helpers
# # ---------------------------------------------------------------------

# def count_merged_cells(excel_path: str, sheet_name=None) -> int:
#     suffix = Path(excel_path).suffix.lower()
#     if suffix == ".xlsx":
#         from openpyxl import load_workbook
#         wb = load_workbook(excel_path, read_only=False, data_only=True)
#         sheets = [sheet_name] if sheet_name else wb.sheetnames
#         total = 0
#         for s in sheets:
#             ws = wb[s]
#             total += len(ws.merged_cells.ranges)
#         wb.close()
#         return total
#     elif suffix == ".xls":
#         import xlrd
#         wb = xlrd.open_workbook(excel_path, formatting_info=True)
#         sheets = [sheet_name] if sheet_name else wb.sheet_names()
#         total = 0
#         for s in sheets:
#             sh = wb.sheet_by_name(s)
#             total += len(getattr(sh, "merged_cells", []))
#         return total
#     return 0

# def calc_text_density(df: pd.DataFrame) -> float:
#     vals = df.values.flatten()
#     text_like = sum(1 for v in vals if isinstance(v, str) and any(c.isalpha() for c in v))
#     return text_like / (len(vals) + 1e-6)

# def calc_col_fluctuation(df: pd.DataFrame) -> float:
#     counts = df.apply(lambda r: sum(str(v).strip() != "" for v in r), axis=1)
#     return counts.std()

# def detect_sheet_type(df, merged_count):
#     nrows, ncols = df.shape
#     text_density = calc_text_density(df)
#     col_fluct = calc_col_fluctuation(df)

#     if merged_count > 2 and nrows < 100:
#         return "form"
#     if ncols <= 2 and text_density > 0.8:
#         return "schema"
#     if nrows > 250:
#         return "large"
#     if col_fluct > 1.5:
#         return "multi-table"
#     return "table"

# # ---------------------------------------------------------------------
# # 🧠 Main Controller
# # ---------------------------------------------------------------------

# def process_excel_for_rag(excel_path: str, out_dir: str, sheet_name=None):
#     out = Path(out_dir)
#     out.mkdir(parents=True, exist_ok=True)

#     xls = pd.ExcelFile(excel_path)
#     sheets = [sheet_name] if sheet_name else xls.sheet_names

#     for s in sheets:
#         print(f"\n📄 Processing sheet: {s}")
#         df = pd.read_excel(xls, sheet_name=s, engine="openpyxl" if excel_path.endswith("x") else None)
#         df = df.fillna("")
#         merged_count = count_merged_cells(excel_path, s)
#         sheet_type = detect_sheet_type(df, merged_count)
#         print(f"Detected type → {sheet_type} (merged={merged_count})")

#         # Route by type
#         if sheet_type in ("schema", "large"):
#             csv_path = out / f"{s}.csv"
#             df.to_csv(csv_path, index=False)
#             print(f"💾 Saved as CSV → {csv_path}")
#         elif sheet_type == "form":
#             # pdf_path = out / f"{s}_form.pdf"
#             # df_to_pdf(df, pdf_path, title=f"{s} (Form-like)")
#             print("normal flow")
#             return "normal flow"
#         elif sheet_type == "multi-table":
#             # split on blank rows
#             groups, current = [], []
#             for _, row in df.iterrows():
#                 if all(str(v).strip() == "" for v in row):
#                     if current:
#                         groups.append(pd.DataFrame(current, columns=df.columns))
#                         current = []
#                 else:
#                     current.append(row)
#             if current:
#                 groups.append(pd.DataFrame(current, columns=df.columns))
#             for i, sub in enumerate(groups, 1):
#                 pdf_path = out / f"{s}_part{i}.pdf"
#                 df_to_pdf(sub, pdf_path, title=f"{s} (part {i})")
#         else:  # table
#             pdf_path = out / f"{s}_table.pdf"
#             df_to_pdf(df, pdf_path, title=s)

# # ---------------------------------------------------------------------
# # ▶️ Example Run
# # ---------------------------------------------------------------------

# if __name__ == "__main__":
#     excel_path = "Haviland-Ni Sulfamate 11-12-24.xls"  # sample file
#     excel_path = "Logsheet.xlsx"
#     excel_path = "columns.xlsx"

#     output_dir = "output"
#     process_excel_for_rag(excel_path, output_dir)


"""
excel_to_rag_vfinal.py
----------------------

Enhanced Excel → RAG preprocessing system.

Features:
- Handles .xls / .xlsx
- Detects forms, tables, multi-tables, large data
- Smart export to CSV or PDF
- Supports page size config (A4 → Tabloid)
- Embeds metadata: filename, sheetname, classification
"""

from pathlib import Path
import pandas as pd
from reportlab.lib.pagesizes import A4, A3, A2, A1, A0, landscape, portrait
from reportlab.lib.pagesizes import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfdoc
from reportlab.lib.units import mm
import numpy as np

# ---------------------------------------------------------------------
# 📄 Page Sizes
# ---------------------------------------------------------------------

PAGE_SIZES = {
    "A4": A4,
    "A3": A3,
    "A2": A2,
    "A1": A1,
    "A0": A0,
    "Tabloid": (11 * inch, 17 * inch),
}

# ---------------------------------------------------------------------
# 🧩 Utility functions
# ---------------------------------------------------------------------

def safe_str(x):
    return "" if pd.isna(x) else str(x).strip()

def mm_to_pt(mm_val):
    return mm_val * 2.83465

# ---------------------------------------------------------------------
# 📏 PDF Helpers
# ---------------------------------------------------------------------

def compute_column_widths_auto(df: pd.DataFrame, page_width_pt: float, margin_pt: float):
    usable_width = page_width_pt - 2 * margin_pt
    ncols = len(df.columns)
    if ncols == 0:
        return []
    weights = []
    for col in df.columns:
        col_len = len(str(col))
        avg_len = min(df[col].astype(str).map(len).mean(), 80)
        weights.append(max(col_len * 1.5, avg_len))
    total_weight = sum(weights)
    widths = [usable_width * (w / total_weight) for w in weights]
    min_width_pt, max_width_pt = 20 * mm, 90 * mm
    widths = [min(max(w, min_width_pt), max_width_pt) for w in widths]
    total = sum(widths)
    if total > usable_width:
        scale = usable_width / total
        widths = [w * scale for w in widths]
    return widths

# ---------------------------------------------------------------------
# 🧾 PDF Table Renderer
# ---------------------------------------------------------------------

def df_to_pdf(df: pd.DataFrame, pdf_path: Path, title=None, orientation="auto",
              page_size="A4", metadata=None):
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    df = df.fillna("")
    ncols = len(df.columns)
    max_cell_len = df.astype(str).applymap(len).max().max()

    # pick page size dynamically if auto
    if page_size.lower() == "auto":
        if ncols > 10 or max_cell_len > 50:
            page_size = "A3"
        elif ncols > 15:
            page_size = "Tabloid"
        else:
            page_size = "A4"

    size = PAGE_SIZES.get(page_size, A4)

    # orientation
    if orientation == "auto":
        orientation = "landscape" if ncols > 8 or max_cell_len > 40 else "portrait"
    page_size_final = landscape(size) if orientation == "landscape" else portrait(size)

    page_width, _ = page_size_final
    margin_pt = mm_to_pt(15)
    styles = getSampleStyleSheet()
    normal_style = ParagraphStyle(
        "wrapped",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        alignment=0,
    )

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=page_size_final,
        leftMargin=margin_pt,
        rightMargin=margin_pt,
        topMargin=mm_to_pt(15),
        bottomMargin=mm_to_pt(15),
        title=metadata.get("title", "") if metadata else "",
        author=metadata.get("author", "Excel Parser Auto") if metadata else "",
        subject=metadata.get("subject", "") if metadata else "",
        keywords=metadata.get("keywords", "") if metadata else "",
        creator="RAG Preprocessor v1.0"
    )

    data = [[Paragraph(f"<b>{safe_str(c)}</b>", normal_style) for c in df.columns]]
    for _, row in df.iterrows():
        data.append([Paragraph(safe_str(v).replace("\n", "<br/>"), normal_style) for v in row])

    col_widths = compute_column_widths_auto(df, page_width, margin_pt)
    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements = []
    if title:
        elements.append(Paragraph(f"<b>{title}</b>", styles["Heading2"]))
        elements.append(Spacer(1, 8))
    elements.append(table)
    doc.build(elements)
    print(f"✅ PDF created: {pdf_path}")

# ---------------------------------------------------------------------
# 📊 Classification Helpers
# ---------------------------------------------------------------------

def count_merged_cells(excel_path: str, sheet_name=None) -> int:
    suffix = Path(excel_path).suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=False, data_only=True)
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        total = 0
        for s in sheets:
            ws = wb[s]
            total += len(ws.merged_cells.ranges)
        wb.close()
        return total
    elif suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(excel_path, formatting_info=True)
        sheets = [sheet_name] if sheet_name else wb.sheet_names()
        total = 0
        for s in sheets:
            sh = wb.sheet_by_name(s)
            total += len(getattr(sh, "merged_cells", []))
        return total
    return 0

def calc_text_density(df: pd.DataFrame) -> float:
    vals = df.values.flatten()
    text_like = sum(1 for v in vals if isinstance(v, str) and any(c.isalpha() for c in v))
    return text_like / (len(vals) + 1e-6)

def calc_col_fluctuation(df: pd.DataFrame) -> float:
    counts = df.apply(lambda r: sum(str(v).strip() != "" for v in r), axis=1)
    return counts.std()

def detect_sheet_type(df, merged_count):
    nrows, ncols = df.shape
    text_density = calc_text_density(df)
    col_fluct = calc_col_fluctuation(df)

    if merged_count > 2 and nrows < 100:
        return "form"
    if ncols <= 2 and text_density > 0.8:
        return "schema"
    if nrows > 250:
        return "large"
    if col_fluct > 1.5:
        return "multi-table"
    return "table"

# ---------------------------------------------------------------------
# 🧠 Main Controller
# ---------------------------------------------------------------------


# from pathlib import Path
# import pandas as pd

# def process_excel_for_rag(excel_path: str, out_dir: str, sheet_name=None, page_size="auto"):
#     excel_path = Path(excel_path)

#     # ✅ Strict file type check
#     if excel_path.suffix.lower() not in (".xls", ".xlsx"):
#         print("only_excel_inputs")
#         return "only_excel_inputs"

#     out = Path(out_dir)
#     out.mkdir(parents=True, exist_ok=True)

#     xls = pd.ExcelFile(excel_path)
#     sheets = [sheet_name] if sheet_name else xls.sheet_names
#     orig_filename = excel_path.name

#     output_files = []

#     for s in sheets:
#         print(f"\n📄 Processing sheet: {s}")

#         # ✅ Fix: detect engine correctly for Path
#         ext = excel_path.suffix.lower()
#         engine = "openpyxl" if ext == ".xlsx" else None

#         df = pd.read_excel(xls, sheet_name=s, engine=engine)
#         df = df.fillna("")

#         merged_count = count_merged_cells(str(excel_path), s)
#         sheet_type = detect_sheet_type(df, merged_count)
#         print(f"Detected type → {sheet_type} (merged={merged_count})")

#         # Metadata for PDF
#         meta = {
#             "title": s,
#             "subject": sheet_type,
#             "keywords": orig_filename,
#             "author": "Excel Parser Auto"
#         }

#         if sheet_type in ("schema", "large"):
#             csv_path = out / f"{s}.csv"
#             df.to_csv(csv_path, index=False)
#             print(f"💾 Saved as CSV → {csv_path}")
#             output_files.append(str(csv_path))

#         elif sheet_type == "form":
#             print("normal flow it will be for this")
#             output_files.append("normal flow")

#         elif sheet_type == "multi-table":
#             groups, current = [], []
#             for _, row in df.iterrows():
#                 if all(str(v).strip() == "" for v in row):
#                     if current:
#                         groups.append(pd.DataFrame(current, columns=df.columns))
#                         current = []
#                 else:
#                     current.append(row)
#             if current:
#                 groups.append(pd.DataFrame(current, columns=df.columns))
#             for i, sub in enumerate(groups, 1):
#                 pdf_path = out / f"{s}_part{i}.pdf"
#                 df_to_pdf(sub, pdf_path, title=f"{s} (part {i})", page_size=page_size, metadata=meta)
#                 print(f"🧾 Saved multi-table PDF → {pdf_path}")
#                 output_files.append(str(pdf_path))

#         else:
#             pdf_path = out / f"{s}_table.pdf"
#             df_to_pdf(df, pdf_path, title=s, page_size=page_size, metadata=meta)
#             print(f"🧾 Saved table PDF → {pdf_path}")
#             output_files.append(str(pdf_path))

#         return output_files

def process_excel_for_rag(excel_path: str, out_dir: str, sheet_name=None, page_size="auto"):
    excel_path = Path(excel_path)

    # ✅ Strict file type check
    if excel_path.suffix.lower() not in (".xls", ".xlsx"):
        print("only_excel_inputs")
        return ["only_excel_inputs"]

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    xls = pd.ExcelFile(excel_path)
    sheets = [sheet_name] if sheet_name else xls.sheet_names
    orig_filename = excel_path.name

    output_files = []

    for s in sheets:
        print(f"\n📄 Processing sheet: {s}")

        ext = excel_path.suffix.lower()
        engine = "openpyxl" if ext == ".xlsx" else None
        df = pd.read_excel(xls, sheet_name=s, engine=engine)
        df = df.fillna("")

        merged_count = count_merged_cells(str(excel_path), s)
        sheet_type = detect_sheet_type(df, merged_count)
        print(f"Detected type → {sheet_type} (merged={merged_count})")

        meta = {
            "title": s,
            "subject": sheet_type,
            "keywords": orig_filename,
            "author": "Excel Parser Auto"
        }

        if sheet_type in ("schema", "large"):
            csv_path = out / f"{s}.csv"
            df.to_csv(csv_path, index=False)
            print(f"💾 Saved as CSV → {csv_path}")
            output_files.append(str(csv_path))

        elif sheet_type == "form":
            print("normal flow it will be for this")
            output_files.append("normal flow")

        elif sheet_type == "multi-table":
            groups, current = [], []
            for _, row in df.iterrows():
                if all(str(v).strip() == "" for v in row):
                    if current:
                        groups.append(pd.DataFrame(current, columns=df.columns))
                        current = []
                else:
                    current.append(row)
            if current:
                groups.append(pd.DataFrame(current, columns=df.columns))

            for i, sub in enumerate(groups, 1):
                pdf_path = out / f"{s}_part{i}.pdf"
                df_to_pdf(sub, pdf_path, title=f"{s} (part {i})", page_size=page_size, metadata=meta)
                print(f"🧾 Saved multi-table PDF → {pdf_path}")
                output_files.append(str(pdf_path))

        else:
            pdf_path = out / f"{s}_table.pdf"
            df_to_pdf(df, pdf_path, title=s, page_size=page_size, metadata=meta)
            print(f"🧾 Saved table PDF → {pdf_path}")
            output_files.append(str(pdf_path))

    return output_files


# ---------------------------------------------------------------------
# # ▶️ Example Run
# # ---------------------------------------------------------------------

if __name__ == "__main__":
    excel_path = "Haviland-Ni Sulfamate 11-12-24.xls"
    excel_path = "../temo/Logsheet.xlsx"
    output_dir = "excel_converted"
    ans = process_excel_for_rag(excel_path, output_dir, page_size="A4")
    print(ans)
